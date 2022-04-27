from operator import contains
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, NamedStyle, Color, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def make_table(df, file, sheet_name):
    # file = 'data/output/bosch wrangled.xlsx'
    writer_args = {
        'path': file,
        'mode': 'w',
        'engine': 'openpyxl'}
        
    # https://openpyxl.readthedocs.io/en/stable/formatting.html#colorscale
    percentile_rule = ColorScaleRule(
        start_type='percentile', start_value=.25, start_color='ffaaaa',  # red-ish
        mid_type='num', mid_value=.8, mid_color='ffffff',  # value zero==white
        end_type='percentile', end_value=1, end_color='aaffaa')  # green-ish

    redFill = PatternFill(start_color='ffaaaa', end_color='ffaaaa', fill_type='solid')
    
    match_rule_NONE = CellIsRule(operator='equal',formula=['"NONE"','"SIMILAR"'], stopIfTrue=False, fill=redFill)  # red-ish
    match_rule_SIMILAR = CellIsRule(operator='equal',formula=['"SIMILAR"'], stopIfTrue=False, fill=redFill)  # red-ish    
        
    # create a custom named style for the Confidence scores
    confidence_style = NamedStyle(
        name="Confidence Style",
        number_format='0.00',
        # font=Font(color='999999', italic=True),
        alignment=Alignment(horizontal='center', vertical='top'))
    
    df.index = np.arange(2, len(df)+2) # shift row (index) to start at 1
    df.index.names = ['Original Row']

    confidence_cols = []
    for i, colm in enumerate(df,2): # start at 2 b/c an index will be added when converts to excel sheet
        if 'Confidence' in colm:
            colm_ltr = get_column_letter(i)
            confidence_cols.append(colm_ltr)

    match_col = []
    for i, colm in enumerate(df,2): # start at 2 b/c an index will be added when converts to excel sheet
        if 'Match' in colm:
            colm_ltr = get_column_letter(i)
            match_col.append(colm_ltr)
   
    tbl_sheet = sheet_name+'_tbl'
    with pd.ExcelWriter(**writer_args) as xlsx:
        df.to_excel(xlsx, tbl_sheet)        
        # worksheets that have been created with this ExcelWriter can be accessed
        # by openpyxl using its API. `ws` is now a openpyxl Worksheet object
        ws = xlsx.sheets[tbl_sheet]

        # set colum widths based on longest values
        factor = 1.05
        for idx, col in enumerate(ws.iter_cols(), 1):
            vals = (len(u"{0}".format(c.value)) for c in col if c.value is not None)
            max_width = max(vals) * factor
            colm_width = min(80,max(max_width, 20))
            ws.column_dimensions[get_column_letter(idx)].width = colm_width

        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=True,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=False)    
        table = Table(displayName=ws.title,ref=ws.dimensions,tableType='worksheet')
        table.tableStyleInfo = style

        ws.add_table(table)

        # format cells (colms) where we want to assign a particular style  
        for row in ws.iter_rows():  
            for cell in row:      
                cell.alignment = Alignment(wrap_text=True,vertical='top')
                if cell.column_letter in confidence_cols:# if cell.data_type == 'n': #  if cell.column_letter == 'G':
                    cell.style = confidence_style

        # conditional formatting, calls color function above
        for colm_ltr in confidence_cols:
            value_cells = colm_ltr+'2:'+colm_ltr+str(ws.max_row)
            ws.conditional_formatting.add(value_cells, percentile_rule)

        for colm_ltr in match_col:
            value_cells = colm_ltr+'2:'+colm_ltr+str(ws.max_row)
            ws.conditional_formatting.add(value_cells, match_rule_NONE)
            ws.conditional_formatting.add(value_cells, match_rule_SIMILAR)


