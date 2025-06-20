import re as _re
from typing import BinaryIO
from ..utils import wildcard_expansion as _wildcard_expansion
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import xlsxwriter
import pandas as _pd
from pandas.api.types import is_scalar
import numpy as _np
import xlsxwriter.utility as _xlu


# ──────────────────────────────────────────────────────────────────────────────
# Module‐level: Default header style (defined exactly once)
# ──────────────────────────────────────────────────────────────────────────────
DEFAULT_HEADER_STYLE = {
    'bold'      : True,
    'align'     : 'center',
    'valign'    : 'vcenter',
    'font_color': '#FFFFFF',   # white
    'fg_color'  : '#4F81BD',    # blue fill
    'text_wrap' : True
}


def _is_na(val):
    """
    Return True only if `val` is a scalar and is NaN/None/NaT.
    Arrays, lists, tuples, Series, etc. are never considered NA here.
    """
    return is_scalar(val) and _pd.isna(val)

# class WranglesXLSXWriter(xlsxwriter):
#     print()


def file_formatting(
    wb, #workbook
    worksheet,
    file_name,
    column_settings,
    sheet_name='Sheet1',
    style_definitions=None,
    column_styles=None,
    group_by=None,
    separator_columns=None,
    conditional_formats=None,
    checkbox_columns=None,
    freeze_panes=None,
    column_widths=None,
    column_formulas=None,
    **kwargs
   ):
    """
    Export a pandas DataFrame to an Excel file with rich formatting via XlsxWriter.

    Parameters:
      df (pd.DataFrame): DataFrame to write.
      file_name (str): Output Excel file path.
      sheet_name (str): Worksheet name (default 'Sheet1').
      style_definitions (dict): {style_name: format_props_dict} for reusable styles.
      column_styles (dict): Map column → style spec. Each spec can be:
          • a string (named style_name from style_definitions),
          • a dict of direct format props (applies to data cells only),
          • or a dict containing any of {'header', 'data', 'both'}:
            header: style_name or inline dict for header only,
            data:   style_name or inline dict for data only,
            both:   style_name or inline dict for both header & data.
      group_by (str or list): Column name(s) to group rows; draws a bottom border when the key changes.
      separator_columns (str, int, list): Column(s) after which to draw a vertical (medium) border.
      conditional_formats (dict): {column: rule}, where rule can be:
          • 'RYG' (built-in red/yellow/green 3-color scale),
          • '3_color_scale', '2_color_scale', 'data_bar', etc.,
          • or a full XlsxWriter rule dict (with type, colors, thresholds).
      checkbox_columns (str or list): Column(s) whose boolean cells become in-cell checkboxes.
      freeze_panes (bool or str or int): If True or str/int, freeze header row plus columns up to and including named column.
          • If True, no columns frozen (just header).
          • If a column name (string), freeze all columns up to that column.
          • If an integer index, freeze up to that index.
      column_widths (dict): {column_name or index: width} for explicit column widths.

    Returns:
      The original DataFrame (in case chaining is desired).
    """
    
    # writer = _pd.ExcelWriter(file_name, engine="xlsxwriter")

    # # Write the dataframe data to XlsxWriter, turn off the default index
    # df.to_excel(writer, sheet_name, index=False)

    # ──────────────────────────────────────────────────────────────────────────
    # 1) BUILD STYLE DEFINITIONS
    # ──────────────────────────────────────────────────────────────────────────

    # Start by injecting our one default header style, then layer on user styles.
    style_def = {'header_default': DEFAULT_HEADER_STYLE}
    if style_definitions:
        style_def.update(style_definitions)

    # Prepare lookup dicts for each column:
    #   data_style_by_col[col]   → style_name for data cells
    #   header_style_by_col[col] → style_name for header cell
    data_style_by_col   = {}
    header_style_by_col = {}
    auto_counter = 0

    def _next_auto_name():
        nonlocal auto_counter
        name = f"__auto_{auto_counter}"
        auto_counter += 1
        return name

    # 1.1) Resolve column_styles into two maps
    # for col in df.columns:
    #     spec = (column_styles or {}).get(col, 'default')

    #     # Case A: A dict that explicitly uses any of 'header','data','both'
    #     if isinstance(spec, dict) and any(k in spec for k in ('header','data','both')):
    #         if 'both' in spec:
    #             data_part   = header_part = spec['both']
    #         else:
    #             data_part   = spec.get('data',   'default')
    #             header_part = spec.get('header', 'header_default')
    #         data_style_by_col[col]   = data_part
    #         header_style_by_col[col] = header_part

    #     # Case B: A simple string or inline-dict → data style only, header gets default
    #     elif isinstance(spec, (str, dict)):
    #         data_style_by_col[col]   = spec
    #         header_style_by_col[col] = 'header_default'

    #     else:
    #         raise TypeError(
    #             f"column_styles[{col!r}] must be a str, dict, or a dict with 'header'/'data'/'both'; "
    #             f"got {type(spec).__name__}"
    #         )

    # 1.2) Turn any inline-dict into a real named style in style_def
    for mapping in (data_style_by_col, header_style_by_col):
        for col, style_spec in mapping.items():
            if isinstance(style_spec, dict):
                auto_name = _next_auto_name()
                style_def[auto_name] = style_spec
                mapping[col] = auto_name
            elif isinstance(style_spec, str):
                # Ensure the named style exists (even if empty); user may override it in style_definitions.
                style_def.setdefault(style_spec, {})
            else:
                raise TypeError(f"Invalid style spec for column {col!r}: {style_spec!r}")

    # 1.3) Ensure a 'default' style exists for any column not explicitly styled
    style_def.setdefault('default', {})

    # ──────────────────────────────────────────────────────────────────────────
    # 2) CREATE WORKBOOK, WORKSHEET, AND FORMAT OBJECTS
    # ──────────────────────────────────────────────────────────────────────────
    # workbook  = writer.book
    formats   = {name: wb.add_format(props) for name, props in style_def.items()}
    # worksheet = df.sheets[sheet_name]

    # Initialize containers that will be used for borders:
    border_after_indices = set()  # columns that require a vertical (medium) border
    bordered_cache = {}           # cache for Format objects combined with bottom/right border

    # ──────────────────────────────────────────────────────────────────────────
    # 3) FREEZE PANES LOGIC
    # ──────────────────────────────────────────────────────────────────────────
    if freeze_panes:
        # If freeze_panes is a column name:
        if isinstance(freeze_panes, str):
            if freeze_panes not in df.columns:
                raise ValueError(f"freeze_panes: column '{freeze_panes}' not found in DataFrame")
            col_idx = df.columns.get_loc(freeze_panes)
        elif isinstance(freeze_panes, int):
            col_idx = freeze_panes
            if not (0 <= col_idx < len(df.columns)):
                raise IndexError(f"freeze_panes index {col_idx} out of range for {len(df.columns)} columns")
        elif freeze_panes is True:
            # Default: no columns frozen, just header row
            worksheet.freeze_panes(1, 0)
            col_idx = None
        else:
            raise TypeError("freeze_panes must be True, a column name (str), or an integer index")

        # If a column name/index was provided, freeze that many columns to the left
        if isinstance(col_idx, int):
            worksheet.freeze_panes(1, col_idx + 1)

    # ──────────────────────────────────────────────────────────────────────────
    # 4) DETERMINE SEPARATOR COLUMNS (vertical borders)
    # # ──────────────────────────────────────────────────────────────────────────
    # if separator_columns:
    #     sep_cols = (
    #         [separator_columns]
    #         if not isinstance(separator_columns, (list, tuple))
    #         else separator_columns
    #     )
    #     for col in sep_cols:
    #         if isinstance(col, str):
    #             if col not in df.columns:
    #                 raise ValueError(f"separator_columns: column '{col}' not found")
    #             idx = df.columns.get_loc(col)
    #         else:
    #             idx = col
    #             if not (0 <= idx < len(df.columns)):
    #                 raise IndexError(f"separator_columns index {idx} out of range")
    #         border_after_indices.add(idx)

    # ──────────────────────────────────────────────────────────────────────────
    # 5) DETERMINE ROW GROUP BREAKPOINTS (horizontal bottom borders)
    # ──────────────────────────────────────────────────────────────────────────
    # group_break_rows = set()
    # if group_by:
    #     group_cols = [group_by] if isinstance(group_by, str) else list(group_by)
    #     for gc in group_cols:
    #         if gc not in df.columns:
    #             raise ValueError(f"group_by: column '{gc}' not found")

    #     group_indices = [df.columns.get_loc(c) for c in group_cols]
    #     for pos in range(len(df) - 1):
    #         for ci in group_indices:
    #             curr = df.iat[pos, ci]
    #             nxt  = df.iat[pos + 1, ci]
    #             if not _pd.isna(curr) and not _pd.isna(nxt):
    #                 if curr != nxt:
    #                     group_break_rows.add(pos + 1)
    #                     break
    #             elif _pd.isna(curr) != _pd.isna(nxt):
    #                 group_break_rows.add(pos + 1)
    #                 break

    # ──────────────────────────────────────────────────────────────────────────
    # 6) WRITE HEADER ROW (ROW 0)
    # ──────────────────────────────────────────────────────────────────────────
    # for col_idx, col_name in enumerate(df.columns):
    #     fmt = formats[header_style_by_col[col_name]]

    #     # If this column is marked as a separator, add a medium right border
    #     if col_idx in border_after_indices:
    #         key = (header_style_by_col[col_name], 'right')
    #         if key not in bordered_cache:
    #             props = style_def[header_style_by_col[col_name]].copy()
    #             props['right'] = 2  # medium border
    #             bordered_cache[key] = workbook.add_format(props)
    #         fmt = bordered_cache[key]

    #     worksheet.write(0, col_idx, col_name, fmt)

    # # ──────────────────────────────────────────────────────────────────────────
    # # 7) WRITE DATA ROWS (ROWS 1 .. N)
    # # ──────────────────────────────────────────────────────────────────────────
    # for row_num, record in enumerate(df.itertuples(index=False, name=None), start=1):
    #     is_group_break = (row_num in group_break_rows)
    #     for col_idx, value in enumerate(record):
    #         col_name   = df.columns[col_idx]
    #         style_name = data_style_by_col[col_name]
    #         fmt        = formats[style_name]

    #         # 7.1) Apply medium bottom border if group boundary
    #         #      and/or medium right border if separator column
    #         need_bottom = is_group_break
    #         need_right  = (col_idx in border_after_indices)
    #         if need_bottom or need_right:
    #             key_flags = (style_name, need_bottom, need_right)
    #             if key_flags not in bordered_cache:
    #                 props = style_def[style_name].copy()
    #                 if need_bottom:
    #                     props['bottom'] = 2
    #                 if need_right:
    #                     props['right'] = 2
    #                 bordered_cache[key_flags] = workbook.add_format(props)
    #             fmt = bordered_cache[key_flags]

    #         # 7.2) Value‐type dispatch
    #         if _is_na(value) or value is None:
    #             worksheet.write_blank(row_num, col_idx, None, fmt)

    #         elif isinstance(value, str):
    #             if value.startswith(('http://', 'https://', 'ftp://', 'mailto:')) or value.startswith('www.'):
    #                 url = value if value.startswith(('http', 'ftp', 'mailto')) else 'http://' + value
    #                 worksheet.write_url(row_num, col_idx, url, fmt, string=value)
    #             else:
    #                 worksheet.write(row_num, col_idx, value, fmt)

    #         elif isinstance(value, bool):
    #             worksheet.write_boolean(row_num, col_idx, value, fmt)

    #         elif isinstance(value, (int, float, _np.number)) and not _np.isnan(value):
    #             worksheet.write_number(row_num, col_idx, float(value), fmt)

    #         elif isinstance(value, dict):
    #             # Dump dict as JSON string
    #             worksheet.write(row_num, col_idx, _pd.io.json.dumps(value, ensure_ascii=False), fmt)

    #         elif isinstance(value, (list, tuple, _np.ndarray, _pd.Series)):
    #             flat = ', '.join(map(str, _np.asarray(value).ravel()))
    #             worksheet.write(row_num, col_idx, flat, fmt)

    #         else:
    #             # Fallback: cast to string
    #             worksheet.write(row_num, col_idx, str(value), fmt)

    # # ──────────────────────────────────────────────────────────────────────────
    # # 8) APPLY CONDITIONAL FORMATTING
    # # ──────────────────────────────────────────────────────────────────────────
    # if conditional_formats:
    #     for col, rule in conditional_formats.items():
    #         if isinstance(col, str):
    #             if col not in df.columns:
    #                 raise ValueError(f"conditional_formats: column '{col}' not found")
    #             col_idx = df.columns.get_loc(col)
    #         else:
    #             col_idx = col

    #         first_data_row = 1
    #         last_data_row  = len(df)

    #         # Expand simple string shortcuts
    #         if isinstance(rule, str):
    #             rl = rule.lower()
    #             if rl in ('ryg', 'red-yellow-green', 'red_yellow_green'):
    #                 options = {
    #                     'type':      '3_color_scale',
    #                     'min_color': '#F8696B',
    #                     'mid_color': '#FFEB84',
    #                     'max_color': '#63BE7B'
    #                 }
    #             elif '3_color' in rl:
    #                 options = {'type': '3_color_scale'}
    #             elif '2_color' in rl:
    #                 options = {'type': '2_color_scale'}
    #             elif 'data_bar' in rl:
    #                 options = {'type': 'data_bar'}
    #             else:
    #                 continue
    #         elif isinstance(rule, dict):
    #             options = rule.copy()
    #         else:
    #             continue

    #         worksheet.conditional_format(first_data_row, col_idx, last_data_row, col_idx, options)

    # # ──────────────────────────────────────────────────────────────────────────
    # # 9) SET COLUMN WIDTHS OR AUTOFIT (with an upper bound of 50)
    # # ──────────────────────────────────────────────────────────────────────────
    # if column_widths:
    #     # 9.1) First, apply any user‐specified widths verbatim
    #     for col, width in column_widths.items():
    #         if isinstance(col, str):
    #             if col not in df.columns:
    #                 raise ValueError(f"column_widths: column '{col}' not found")
    #             idx = df.columns.get_loc(col)
    #         else:
    #             idx = col
    #             if not (0 <= idx < len(df.columns)):
    #                 raise IndexError(f"column_widths index {idx} out of range")
    #         worksheet.set_column(idx, idx, width)

    #     # 9.2) Now auto‐fit any columns NOT in column_widths,
    #     #       capping the final width at 50.
    #     #       If the user’s explicit width is > 50, we leave it as is;
    #     #       otherwise, we compute max_len+2 and cap at 50.
    #     for idx, col in enumerate(df.columns):
    #         # Skip columns that already have an explicit width
    #         if ( (isinstance(col, str) and col in column_widths)
    #           or (isinstance(idx, int) and idx in column_widths) ):
    #             continue

    #         # Compute the “ideal” width from data+header
    #         values = df[col]
    #         max_len = max([len(str(x)) for x in values] + [len(str(col))])
    #         ideal_width = max_len + 2

    #         # Cap at 50
    #         final_width = ideal_width if ideal_width <= 50 else 50
    #         worksheet.set_column(idx, idx, final_width)

    # else:
    #     # No explicit widths at all → auto‐fit every column with max = 50
    #     for idx, col in enumerate(df.columns):
    #         values = df[col]
    #         max_len = max([len(str(x)) for x in values] + [len(str(col))])
    #         ideal_width = max_len + 2
    #         final_width = ideal_width if ideal_width <= 50 else 50
    #         worksheet.set_column(idx, idx, final_width)

    # ──────────────────────────────────────────────────────────────────────────
    # 10) ADD AN EXCEL TABLE FOR BANDED ROWS, INJECT ANY USER FORMULAS, AND
    #     ENSURE HEADER WRAP (text_wrap=True)
    # ──────────────────────────────────────────────────────────────────────────
    first_row, first_col = 0, 0
    # last_row  = len(df)
    # last_col  = len(df.columns) - 1
    table_range = _xlu.xl_range_abs(first_row, first_col, kwargs['max_row'], kwargs['max_col']-1)

    # 10.1) Build a map from column_name → the header Format object (which has text_wrap=True)
    # wrapped_header_cells = {
    #     col_name: formats[ header_style_by_col[col_name] ]
    #     for col_name in df.columns
    # }

    # # 10.2) Build the “columns” list for add_table, using 'header_format' 
    # columns_list = []
    # for col_name in df.columns:
    #     hdr_fmt  = wrapped_header_cells[col_name]
    #     data_fmt = formats[ data_style_by_col[col_name] ]  # whatever data style you chose

    #     if column_formulas and col_name in column_formulas:
    #         # Formula + wrapping on header + normal data format on each data cell
    #         columns_list.append({
    #             'header'        : col_name,
    #             'formula'       : column_formulas[col_name],
    #             'header_format' : hdr_fmt,   # wrap header
    #             'format'        : data_fmt   # style data cells normally
    #         })
    #     else:
    #         # No formula; still wrap header, and style data cells
    #         columns_list.append({
    #             'header'        : col_name,
    #             'header_format' : hdr_fmt,   # wrap header
    #             'format'        : data_fmt   # style data cells normally
    #         })

    format_dict = {}

    # Add formatting to workbook
    for setting in column_settings:
        format_dict[setting['header']] = wb.add_format(setting.pop('data', {}))
        # setting['format'] = setting['header']

    header_format = wb.add_format({'font_size': 41})
    data_format = wb.add_format({'font_size': 4})

    for setting in column_settings:
        setting['format'] = data_format
        # setting['format'] = format_dict.get(setting['header'], {})
        setting['header_format'] = header_format


    # 10.3) Finally, add the table. Now the header row will use hdr_fmt (with text_wrap),
    # and the data cells will use data_fmt (e.g. your checkbox or green‐font format).
    worksheet.add_table(table_range, {
        'name'         : sheet_name.replace(' ', '_'),
        # 'columns'      : kwargs['column_formats'],
        'columns'      : column_settings,
        'style'        : 'Table Style Medium9',
        'banded_rows'  : (group_by is None)
    })

    # ──────────────────────────────────────────────────────────────────────────
    # 11) SAVE AND RETURN
    # ──────────────────────────────────────────────────────────────────────────
    wb.close()
    return wb
    # writer.close()
    # return df