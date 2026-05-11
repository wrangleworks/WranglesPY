import pandas as pd
import os
from openai import OpenAI
import concurrent.futures
from tqdm import tqdm
from pydantic import BaseModel, Field
from typing import Literal
from tenacity import retry, wait_random_exponential, stop_after_attempt
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# ===========================================
# PURPOSE:
# Eeads the same workbook and taxonomy structure as the synchronous script.
# Converts each row into a raw /v1/responses request line for batch processing.
# Automatically splits requests into multiple .jsonl files at 50,000 rows per file and also keeps a safety buffer under the 200 MB file limit.
# Uses Structured Outputs for Responses via text.format + JSON schema, which is the correct raw request shape for Responses; the SDK helper responses.parse(..., text_format=...) is convenient for synchronous calls, but batch files need the underlying endpoint parameters.
# Uploads the JSONL files, creates batch jobs, polls status, downloads the result files, parses the JSON output, and rebuilds the final Excel output in roughly the same format as the synchronous script.
# ==========================================

# ==========================================
# CONFIGURATION
# ==========================================
INPUT_FILE_PATH = "./data/Motion/Items and PGCs April 2026 AI classification.xlsx"
OUTPUT_FILE_PATH = "./data/Motion/Classified_Parts_Output.xlsx"

# Hardcoded your exact sheet names
ITEMS_SHEET_NAME = "Item List"   
PGC_SHEET_NAME = "PG Codes"      

# Set to an integer (e.g., 50) to test on a small batch. Set to None to run the whole file.
SAMPLE_SIZE = 200
MAX_CONCURRENT_CALLS = 100 
# ==========================================

# 1. Initialize the client
api_key = os.environ.get("MOTION_OPENAI_API_KEY")
if not api_key:
    raise RuntimeError("Environment variable MOTION_OPENAI_API_KEY not set.")

client = OpenAI(api_key=api_key)

# 2. Load the Taxonomy dynamically from the Excel file
print(f"Loading taxonomy from {INPUT_FILE_PATH} (Sheet: {PGC_SHEET_NAME})...")
taxonomy_df = pd.read_excel(INPUT_FILE_PATH, sheet_name=PGC_SHEET_NAME)

# Ensure the new explicit column headers exist (Added Major and Major_Code)
required_columns = {'Full_Code', 'Major_Code', 'Sub_Code', 'Family_Code', 'Major', 'Sub', 'Family'}
if not required_columns.issubset(taxonomy_df.columns):
    raise ValueError(f"The PG Codes sheet is missing one or more required columns: {required_columns}")

# Create the taxonomy string for the AI prompt (Using the text columns)
taxonomy_df['Predicted_Taxonomy'] = taxonomy_df['Sub'].astype(str) + " > " + taxonomy_df['Family'].astype(str)
CATEGORY_PAIRS = taxonomy_df['Predicted_Taxonomy'].dropna().unique().tolist()

# Add the fallback option
CATEGORY_PAIRS.append("Unknown > Unknown")

from typing import Literal
from pydantic import BaseModel, Field
from tenacity import retry, wait_random_exponential, stop_after_attempt

# 3. Define the Pydantic Schema
class ClassificationResult(BaseModel):
    Predicted_Taxonomy: str = Field(
        description="The exact 'Sub > Family' text pair chosen from the approved category list or recommended as a new pairing."
    )
    Confidence_Score: Literal["High", "Medium", "Low", "N/A"] = Field(
        description="Confidence level of the classification. 'High' if it clearly fits an existing category, 'Medium' if it somewhat fits, 'Low' if it's a poor fit, and 'N/A' if the item is unrecognizable."
    )
    Reasoning: str = Field(
        description="2 sentence explanation. First sentence summarize the parts key descriptive elements. Second sentence explain the reasoning for the classification."
    )
    Proposed_Taxonomy: str = Field(
        description="The proposed new 'Sub > Family' text pair if the classification confidence is Low or a new category is needed."
    )

# 4. Inject the pairs directly into the System Prompt for auto-caching
categories_list_text = "\n".join([f"- {pair}" for pair in CATEGORY_PAIRS])

FEW_SHOT_EXAMPLES = """
# Examples

<example id="1">
<input>
Part Number: KQ2H01-03A
Description: SMC ONE-TOUCH MALE CONNECTOR, 1/8 NPT, 5/32 TUBE
</input>
<output>
{
  "Predicted_Taxonomy": "Pneumatics > Fittings & Tubing",
  "Confidence_Score": "High",
  "Reasoning": "SMC KQ2H01-03A is a pneumatic one-touch male connector for tube-to-thread air line connections. It is a direct fit for Pneumatics > Fittings & Tubing.",
  "Proposed_Taxonomy": ""
}
</output>
</example>

<example id="2">
<input>
Part Number: 2017-MOTOR-3
Description: 121-13624-13 MOTOR W/ 2048 LINE ENCODER & REAR 0.25IN SHAFT
</input>
<output>
{
  "Predicted_Taxonomy": "Electric Motion > Servo Motors",
  "Confidence_Score": "Medium",
  "Reasoning": "This item is a motor assembly with an integrated 2048-line encoder and output shaft, which are characteristic of closed-loop motion control hardware. The encoder strongly suggests a servo-style application, but the description does not explicitly confirm servo functionality, so the fit is somewhat ambiguous.",
  "Proposed_Taxonomy": ""
}
</output>
</example>

<example id="3">
<input>
Part Number: 69915K54
Description: MCMASTER LIQUID TIGHT SEAL 1/2
</input>
<output>
{
  "Predicted_Taxonomy": "Pumps > Pump Repair Parts",
  "Confidence_Score": "Low",
  "Reasoning": "This item is a liquid-tight seal used to protect conduit or cable entry points from moisture and debris ingress. The closest approved category is still a poor fit, so a new taxonomy is warranted.",
  "Proposed_Taxonomy": "Electrical Accessories > Liquid Tight Seals"
}
</output>
</example>

<example id="4">
<input>
Part Number: 2293682
Description: PC 3682
</input>
<output>
{
  "Predicted_Taxonomy": "Unknown > Unknown",
  "Confidence_Score": "N/A",
  "Reasoning": "The description appears to be only a shorthand identifier or internal code. There is not enough descriptive information to determine what the item is.",
  "Proposed_Taxonomy": ""
}
</output>
</example>
"""

DEVELOPER_PROMPT = f"""You are an expert industrial parts classifier specializing in automation, pneumatics, motion control, electrical, and MRO items.
Your task is to classify a list of parts based on their "Part Number" and "Item Description" either into ONE of the approved taxonomy pairs, and 
if the confidence of the classification is low also propose a new taxonomy pair.

APPROVED TAXONOMY PAIRS:
{categories_list_text}

CRITICAL RULES:
1. You MUST select an exact pairing from the list above and output it to the `Predicted_Taxonomy` field.
2. Assign confidence levels as follows:
   - "High" if the part clearly fits an approved category pair.
   - "Medium" if the part fits an approved category pair but there is some ambiguity.
   - "Low" if the part is recognizable and but the fit is poor even with the most relevant approved category pair.
   - "N/A" if the part is unrecognizable or highly ambiguous. In this case, set `Predicted_Taxonomy` to "Unknown > Unknown".
3. If you designate the Confidence_Score as "Low", propose a new taxonomy pair. Otherwise leave this output blank.
4. Proposed taxonomy pairs should follow the same "Sub > Family" category format, and be specified at the same level of generalization as the approved pairs.
5. When generating a new category pairing, reuse one of the upper-level ("Sub") categories from the approved list if one fits. Otherwise, create new values for the both parts of the category pair.
6. Only use "Other Revenue > All Other Revenue" for service charges and fees.  Do not use it for parts.
7. Recognize and apply standard industrial abbreviations (e.g., 'CYL' = Cylinder, 'BRG' = Bearing).

{FEW_SHOT_EXAMPLES}
"""

@retry(wait=wait_random_exponential(min=1, max=10), stop=stop_after_attempt(5))
def call_openai_api(user_prompt: str) -> ClassificationResult:
    response = client.responses.parse(
        model="gpt-5.4-mini",
        reasoning={"effort": "low"},
        instructions=DEVELOPER_PROMPT,
        input=user_prompt,
        text_format=ClassificationResult,
        max_output_tokens=400,
    )
    return response.output_parsed

# 6. The worker function
def classify_row(row):
    item_id = str(row.get("ID", ""))

    part_number = str(row.get("Item ID", "None provided"))
    description = str(row.get("Item Description", "None provided"))

    user_prompt = f"Part Number: {part_number}\nDescription: {description}"

    try:
        result = call_openai_api(user_prompt)

        return {
            "ID": item_id,
            "Predicted_Taxonomy": result.Predicted_Taxonomy,
            "Confidence_Score": result.Confidence_Score,
            "Reasoning": result.Reasoning,
            "Proposed_Taxonomy": result.Proposed_Taxonomy
        }

    except Exception as e:
        return {
            "ID": item_id,
            "Predicted_Taxonomy": "Unknown > Unknown",
            "Confidence_Score": "N/A",
            "Reasoning": f"API error: {e}",
            "Proposed_Taxonomy": ""
        }
        

# 7. Main Execution Block
if __name__ == "__main__":
    
    print(f"Loading items from {INPUT_FILE_PATH} (Sheet: {ITEMS_SHEET_NAME})...")
    df = pd.read_excel(INPUT_FILE_PATH, sheet_name=ITEMS_SHEET_NAME)
    
    if 'ID' in df.columns:
        df['ID'] = df['ID'].astype(str)
        
    if SAMPLE_SIZE:
        print(f"--- RUNNING IN TEST MODE: Randomly sampling {SAMPLE_SIZE} rows ---")
        # Uses random_state=42 so you get the same random mix every time you test.
        df = df.sample(n=SAMPLE_SIZE, random_state=42).copy()
        
        rows_to_process = df.to_dict(orient='records')
        results = []
    
    print(f"Starting synchronous classification using {MAX_CONCURRENT_CALLS} concurrent workers...")
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_CONCURRENT_CALLS) as executor:
        results = list(tqdm(executor.map(classify_row, rows_to_process), total=len(rows_to_process)))
        
    results_df = pd.DataFrame(results)
    
    # Split the predicted string into Sub and Family for human readability
    results_df[['Sub_Category', 'Family_Category']] = results_df['Predicted_Taxonomy'].str.split(' > ', n=1, expand=True)
    
    # Merge results back with the original items DataFrame
    final_df = pd.merge(df, results_df, on="ID", how="left")
    
    # --- RELATIONAL JOIN FOR ALL TEXT AND CODES ---
    # We grab the Major text column, plus all 4 code columns
    lookup_df = taxonomy_df[['Predicted_Taxonomy', 'Major', 'Major_Code', 'Sub_Code', 'Family_Code', 'Full_Code']].drop_duplicates()
    
    # Rename 'Major' to 'Major_Category' for consistency in the output
    lookup_df = lookup_df.rename(columns={'Major': 'Major_Category'})
    
    final_df = pd.merge(final_df, lookup_df, on="Predicted_Taxonomy", how="left")
    
    # Drop the temporary concatenated column
    final_df = final_df.drop(columns=['Predicted_Taxonomy'])
    
    # Handle the "Unknown" fallbacks
    final_df['Major_Category'] = final_df['Major_Category'].fillna("Unknown")
    final_df['Major_Code'] = final_df['Major_Code'].fillna("UNK")
    final_df['Sub_Code'] = final_df['Sub_Code'].fillna("UNK")
    final_df['Family_Code'] = final_df['Family_Code'].fillna("UNK")
    final_df['Full_Code'] = final_df['Full_Code'].fillna("UNK-00")
    final_df['Proposed_Taxonomy'] = final_df['Proposed_Taxonomy'].fillna("")
    
    # --- DYNAMIC COLUMN REORDERING ---
    original_cols = list(df.columns)
    
    if 'Item Description' in original_cols:
        insert_idx = original_cols.index('Item Description') + 1
    else:
        insert_idx = len(original_cols)
        
    # Build the final column order (Major -> Sub -> Family -> Reason ... Codes at end)
    new_col_order = (
        original_cols[:insert_idx] + 
        ['Major_Category', 'Sub_Category', 'Family_Category', 'Confidence_Score', 'Reasoning', 'Proposed_Taxonomy'] + 
        original_cols[insert_idx:] + 
        ['Major_Code', 'Sub_Code', 'Family_Code', 'Full_Code']
    )
    final_df = final_df[new_col_order]

    # Ensure output directory exists
    os.makedirs(os.path.dirname(OUTPUT_FILE_PATH), exist_ok=True)
    
    print(f"Formatting and saving results to {OUTPUT_FILE_PATH}...")
    
    # Write to Excel with custom formatting
    with pd.ExcelWriter(OUTPUT_FILE_PATH, engine='openpyxl') as writer:
        final_df.to_excel(writer, index=False, sheet_name='Classifications')
        worksheet = writer.sheets['Classifications']
        
        for idx, col in enumerate(final_df.columns):
            col_letter = get_column_letter(idx + 1)
            
            # Apply widths
            if col == 'Reasoning':
                worksheet.column_dimensions[col_letter].width = 100 
            else:
                max_len = max(final_df[col].astype(str).map(len).max(), len(col))
                adjusted_width = min(max_len + 2, 60) 
                worksheet.column_dimensions[col_letter].width = adjusted_width
            
            # Apply Vertical Alignment Top and Text Wrap
            for cell in worksheet[col_letter]:
                if col == 'Reasoning':
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical='top')
            
        # Add Excel Table Formatting
        max_row = len(final_df) + 1
        max_col_letter = get_column_letter(len(final_df.columns))
        table_range = f"A1:{max_col_letter}{max_row}"
        
        tab = Table(displayName="ClassificationData", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        worksheet.add_table(tab)
    
    print(f"\n✅ Processing complete! Results exported successfully.")