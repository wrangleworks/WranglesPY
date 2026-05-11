import argparse
import json
import math
import os
import re
import time
from pathlib import Path
from typing import Literal

import pandas as pd
from openai import OpenAI
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pydantic import BaseModel, Field
from pathlib import Path

# ==========================================
# CONFIGURATION
# ==========================================


BASE_DIR = Path(__file__).resolve().parent
# Prefer the workspace `data/Motion` sibling under the repo root, with a
# fallback to a `data/Motion` directory under the current working dir.
DATA_DIR = (BASE_DIR.parent.parent / "data" / "Motion").resolve()
if not DATA_DIR.exists():
    alt = (Path.cwd() / "data" / "Motion").resolve()
    if alt.exists():
        DATA_DIR = alt
    else:
        raise FileNotFoundError(
            f"DATA_DIR not found at {BASE_DIR.parent.parent / 'data' / 'Motion'} or {alt}"
        )

INPUT_FILE_PATH = DATA_DIR / "Items and PGCs April 2026 AI classification.xlsx"
OUTPUT_FILE_PATH = DATA_DIR / "Classified_Parts_Batch_Output_R3.xlsx"

BATCH_WORK_DIR = DATA_DIR / "batch_jobs"

MANIFEST_PATH = BATCH_WORK_DIR / "batch_manifest.json"
BATCH_INPUT_DIR = BATCH_WORK_DIR / "inputs"
BATCH_OUTPUT_DIR = BATCH_WORK_DIR / "outputs"
BATCH_ERROR_DIR = BATCH_WORK_DIR / "errors"

ITEMS_SHEET_NAME = "Item List"
PGC_SHEET_NAME = "PG Codes"

# Set to an integer to test on a smaller sample, or None to process all rows.
SAMPLE_SIZE = None

MODEL_NAME = "gpt-5.4-mini"
REASONING_EFFORT = "low"
MAX_OUTPUT_TOKENS = 400
MAX_REQUESTS_PER_FILE = 50000
MAX_BATCH_FILE_BYTES = 190 * 1024 * 1024  # stay safely below the 200 MB hard limit
COMPLETION_WINDOW = "24h"
POLL_SECONDS = 30
# ==========================================


class ClassificationResult(BaseModel):
    Predicted_Taxonomy: str = Field(
        description="The exact 'Sub > Family' text pair chosen from the approved category list."
    )
    Confidence_Score: Literal["High", "Medium", "Low", "N/A"] = Field(
        description=(
            "Confidence level of the classification. 'High' if it clearly fits an existing "
            "category, 'Medium' if it somewhat fits, 'Low' if the best approved fit is poor, "
            "and 'N/A' if the item is unrecognizable."
        )
    )
    Reasoning: str = Field(
        description=(
            "2 sentence explanation. First sentence summarize the part's key descriptive "
            "elements. Second sentence explain the reasoning for the classification."
        )
    )
    Proposed_Taxonomy: str = Field(
        description=(
            "The proposed new 'Sub > Family' text pair if the classification confidence is Low "
            "or a new category is needed. Otherwise leave blank."
        )
    )


FEW_SHOT_EXAMPLES = """
# Examples

<example id="1">
<input>
Part Number: KQ2H01-03A
Description: SMC ONE-TOUCH MALE CONNECTOR, 1/8 NPT, 5/32 TUBE
</input>
<output>
{
  "Predicted_Taxonomy": "Pneumatics > Fittings & Tubing",
  "Confidence_Score": "High",
  "Reasoning": "SMC KQ2H01-03A is a pneumatic one-touch male connector for tube-to-thread air line connections. It is a direct fit for Pneumatics > Fittings & Tubing.",
  "Proposed_Taxonomy": ""
}
</output>
</example>

<example id="2">
<input>
Part Number: 2017-MOTOR-3
Description: 121-13624-13 MOTOR W/ 2048 LINE ENCODER & REAR 0.25IN SHAFT
</input>
<output>
{
  "Predicted_Taxonomy": "Electric Motion > Servo Motors",
  "Confidence_Score": "Medium",
  "Reasoning": "This item is a motor assembly with an integrated 2048-line encoder and output shaft, which are characteristic of closed-loop motion control hardware. The encoder strongly suggests a servo-style application, but the description does not explicitly confirm servo functionality, so the fit is somewhat ambiguous.",
  "Proposed_Taxonomy": ""
}
</output>
</example>

<example id="3">
<input>
Part Number: 69915K54
Description: MCMASTER LIQUID TIGHT SEAL 1/2
</input>
<output>
{
  "Predicted_Taxonomy": "Pumps > Pump Repair Parts",
  "Confidence_Score": "Low",
  "Reasoning": "This item is a liquid-tight seal used to protect conduit or cable entry points from moisture and debris ingress. The closest approved category is still a poor fit, so a new taxonomy is warranted.",
  "Proposed_Taxonomy": "Electrical Accessories > Liquid Tight Seals"
}
</output>
</example>

<example id="4">
<input>
Part Number: 2293682
Description: PC 3682
</input>
<output>
{
  "Predicted_Taxonomy": "Unknown > Unknown",
  "Confidence_Score": "N/A",
  "Reasoning": "The description appears to be only a shorthand identifier or internal code. There is not enough descriptive information to determine what the item is.",
  "Proposed_Taxonomy": ""
}
</output>
</example>
"""


def ensure_dirs() -> None:
    BATCH_WORK_DIR.mkdir(parents=True, exist_ok=True)
    BATCH_INPUT_DIR.mkdir(parents=True, exist_ok=True)
    BATCH_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    BATCH_ERROR_DIR.mkdir(parents=True, exist_ok=True)


def get_client() -> OpenAI:
    api_key = os.environ.get("MOTION_OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("Environment variable MOTION_OPENAI_API_KEY not set.")
    return OpenAI(api_key=api_key)


def load_taxonomy_and_items() -> tuple[pd.DataFrame, pd.DataFrame, list[str], str]:
    taxonomy_df = pd.read_excel(INPUT_FILE_PATH, sheet_name=PGC_SHEET_NAME)

    required_columns = {
        "Full_Code",
        "Major_Code",
        "Sub_Code",
        "Family_Code",
        "Major",
        "Sub",
        "Family",
    }
    if not required_columns.issubset(taxonomy_df.columns):
        raise ValueError(
            f"The PG Codes sheet is missing one or more required columns: {required_columns}"
        )

    taxonomy_df["Predicted_Taxonomy"] = (
        taxonomy_df["Sub"].astype(str) + " > " + taxonomy_df["Family"].astype(str)
    )
    category_pairs = taxonomy_df["Predicted_Taxonomy"].dropna().unique().tolist()
    if "Unknown > Unknown" not in category_pairs:
        category_pairs.append("Unknown > Unknown")

    categories_list_text = "\n".join(f"- {pair}" for pair in category_pairs)
    
    developer_prompt = f"""You are an expert industrial parts classifier specializing in automation, pneumatics, motion control, electrical, and MRO items.
Your task is to classify a list of parts based on their "Part Number" and "Item Description" either into ONE of the approved taxonomy pairs, and 
if the confidence of the classification is low also propose a new taxonomy pair.

APPROVED TAXONOMY PAIRS:
{categories_list_text}

CRITICAL RULES:
1. You MUST select an exact pairing from the list above and output it to the `Predicted_Taxonomy` field.
2. Assign confidence levels as follows:
   - "High" if the part clearly fits an approved category pair.
   - "Medium" if the part fits an approved category pair but there is some ambiguity.
   - "Low" if the part is recognizable and but the fit is poor even with the most relevant approved category pair.
   - "N/A" if the part is unrecognizable or highly ambiguous. In this case, set `Predicted_Taxonomy` to "Unknown > Unknown".
3. If you designate the Confidence_Score as "Low", propose a new taxonomy pair. Otherwise leave this output blank.
4. Proposed taxonomy pairs should follow the same "Sub > Family" category format, and be specified at the same level of generalization as the approved pairs.
5. When generating a new category pairing, reuse one of the upper-level ("Sub") categories from the approved list if one fits. Otherwise, create new values for the both parts of the category pair.
6. Only use "Other Revenue > All Other Revenue" for service charges and fees. Do not use it for parts.
7. Recognize and apply standard industrial abbreviations (e.g., 'CYL' = Cylinder, 'BRG' = Bearing).
8. Return JSON that matches the requested schema exactly.

{FEW_SHOT_EXAMPLES}
"""

    items_df = pd.read_excel(INPUT_FILE_PATH, sheet_name=ITEMS_SHEET_NAME)
    if "ID" in items_df.columns:
        items_df["ID"] = items_df["ID"].astype(str)
    items_df = items_df.reset_index(drop=True).copy()
    items_df["__row_number"] = items_df.index

    if SAMPLE_SIZE:
        items_df = items_df.head(SAMPLE_SIZE).copy()

    return taxonomy_df, items_df, category_pairs, developer_prompt


def build_response_format() -> dict:
    schema = ClassificationResult.model_json_schema()
    schema["additionalProperties"] = False
    return {
        "type": "json_schema",
        "name": "classification_result",
        "strict": True,
        "schema": schema,
    }


def build_user_prompt(row: pd.Series) -> str:
    part_number = str(row.get("Item ID", "None provided"))
    description = str(row.get("Item Description", "None provided"))
    return f"Part Number: {part_number}\nDescription: {description}"


def sanitize_custom_id_value(value: str) -> str:
    text = str(value).strip() if value is not None else ""
    if not text:
        return "none"
    text = re.sub(r"[^A-Za-z0-9._-]+", "-", text)
    return text[:80]


def build_batch_request(row: pd.Series, developer_prompt: str, response_format: dict) -> dict:
    row_number = int(row["__row_number"])
    item_id = sanitize_custom_id_value(row.get("ID", ""))

    return {
        "custom_id": f"row-{row_number}__id-{item_id}",
        "method": "POST",
        "url": "/v1/responses",
        "body": {
            "model": MODEL_NAME,
            "reasoning": {"effort": REASONING_EFFORT},
            "instructions": developer_prompt,
            "input": build_user_prompt(row),
            "max_output_tokens": MAX_OUTPUT_TOKENS,
            "text": {
                "format": response_format,
            },
        },
    }


def write_jsonl_shards(items_df: pd.DataFrame, developer_prompt: str) -> dict:
    ensure_dirs()
    response_format = build_response_format()

    for existing in Path(BATCH_INPUT_DIR).glob("*.jsonl"):
        existing.unlink()

    files = []
    shard_index = 1
    shard_rows = 0
    shard_bytes = 0
    output_path = Path(BATCH_INPUT_DIR) / f"classification_batch_{shard_index:03d}.jsonl"
    handle = output_path.open("w", encoding="utf-8")

    try:
        for _, row in items_df.iterrows():
            request_obj = build_batch_request(row, developer_prompt, response_format)
            line = json.dumps(request_obj, ensure_ascii=False)
            encoded_len = len((line + "\n").encode("utf-8"))

            rollover = (
                shard_rows >= MAX_REQUESTS_PER_FILE
                or (shard_rows > 0 and shard_bytes + encoded_len > MAX_BATCH_FILE_BYTES)
            )
            if rollover:
                handle.close()
                files.append(
                    {
                        "file_index": shard_index,
                        "path": str(output_path),
                        "request_count": shard_rows,
                        "bytes": shard_bytes,
                    }
                )
                shard_index += 1
                shard_rows = 0
                shard_bytes = 0
                output_path = Path(BATCH_INPUT_DIR) / f"classification_batch_{shard_index:03d}.jsonl"
                handle = output_path.open("w", encoding="utf-8")

            handle.write(line + "\n")
            shard_rows += 1
            shard_bytes += encoded_len

        handle.close()
        files.append(
            {
                "file_index": shard_index,
                "path": str(output_path),
                "request_count": shard_rows,
                "bytes": shard_bytes,
            }
        )
    finally:
        if not handle.closed:
            handle.close()

    manifest = {
        "created_at": int(time.time()),
        "model": MODEL_NAME,
        "endpoint": "/v1/responses",
        "input_file_path": str(INPUT_FILE_PATH),
        "output_file_path": str(OUTPUT_FILE_PATH),
        "sample_size": SAMPLE_SIZE,
        "files": files,
        "batches": [],
    }

    with open(MANIFEST_PATH, "w", encoding="utf-8") as fh:
        json.dump(manifest, fh, indent=2, default=str)

    return manifest


def load_manifest() -> dict:
    if not os.path.exists(MANIFEST_PATH):
        raise FileNotFoundError(
            f"Manifest not found at {MANIFEST_PATH}. Run the prepare step first."
        )
    with open(MANIFEST_PATH, "r", encoding="utf-8") as fh:
        return json.load(fh)


def save_manifest(manifest: dict) -> None:
    with open(MANIFEST_PATH, "w", encoding="utf-8") as fh:
        json.dump(manifest, fh, indent=2)


def submit_batches(client: OpenAI, manifest: dict) -> dict:
    existing_by_path = {batch["local_input_path"]: batch for batch in manifest.get("batches", [])}
    submitted_batches = []

    for file_info in manifest["files"]:
        local_input_path = file_info["path"]
        if local_input_path in existing_by_path:
            submitted_batches.append(existing_by_path[local_input_path])
            continue

        with open(local_input_path, "rb") as fh:
            upload = client.files.create(file=fh, purpose="batch")

        batch = client.batches.create(
            input_file_id=upload.id,
            endpoint="/v1/responses",
            completion_window=COMPLETION_WINDOW,
            metadata={
                "job_type": "motion_taxonomy_classification",
                "input_filename": os.path.basename(local_input_path),
            },
        )

        batch_info = {
            "local_input_path": local_input_path,
            "input_file_id": upload.id,
            "batch_id": batch.id,
            "status": batch.status,
            "output_file_id": getattr(batch, "output_file_id", None),
            "error_file_id": getattr(batch, "error_file_id", None),
        }
        submitted_batches.append(batch_info)

    manifest["batches"] = submitted_batches
    save_manifest(manifest)
    return manifest


def refresh_batch_statuses(client: OpenAI, manifest: dict) -> dict:
    for batch_info in manifest.get("batches", []):
        batch = client.batches.retrieve(batch_info["batch_id"])
        batch_info["status"] = batch.status
        batch_info["output_file_id"] = getattr(batch, "output_file_id", None)
        batch_info["error_file_id"] = getattr(batch, "error_file_id", None)
        batch_info["request_counts"] = (
            batch.request_counts.to_dict() if getattr(batch, "request_counts", None) else None
        )
        batch_info["usage"] = getattr(batch, "usage", None)
    save_manifest(manifest)
    return manifest


def wait_for_batches(client: OpenAI, manifest: dict) -> dict:
    terminal_statuses = {"completed", "failed", "expired", "cancelled"}
    while True:
        manifest = refresh_batch_statuses(client, manifest)
        statuses = [batch["status"] for batch in manifest.get("batches", [])]
        print("Current batch statuses:", ", ".join(statuses))
        if statuses and all(status in terminal_statuses for status in statuses):
            return manifest
        time.sleep(POLL_SECONDS)


def download_file_content(client: OpenAI, file_id: str, dest_path: str) -> None:
    response = client.files.content(file_id)
    content = response.read()
    with open(dest_path, "wb") as fh:
        fh.write(content)


def download_batch_files(client: OpenAI, manifest: dict) -> dict:
    for batch_info in manifest.get("batches", []):
        if batch_info.get("status") != "completed":
            continue

        batch_id = batch_info["batch_id"]
        output_file_id = batch_info.get("output_file_id")
        error_file_id = batch_info.get("error_file_id")

        if output_file_id:
            output_path = os.path.join(BATCH_OUTPUT_DIR, f"{batch_id}_output.jsonl")
            if not os.path.exists(output_path):
                download_file_content(client, output_file_id, output_path)
            batch_info["local_output_path"] = output_path

        if error_file_id:
            error_path = os.path.join(BATCH_ERROR_DIR, f"{batch_id}_errors.jsonl")
            if not os.path.exists(error_path):
                download_file_content(client, error_file_id, error_path)
            batch_info["local_error_path"] = error_path

    save_manifest(manifest)
    return manifest


def extract_output_text(response_body: dict) -> str:
    if isinstance(response_body.get("output_text"), str) and response_body["output_text"].strip():
        return response_body["output_text"]

    chunks = []
    for item in response_body.get("output", []):
        if item.get("type") != "message":
            continue
        for content in item.get("content", []):
            if content.get("type") in {"output_text", "text"}:
                text = content.get("text", "")
                if text:
                    chunks.append(text)
    return "\n".join(chunks).strip()


def parse_custom_id(custom_id: str) -> int:
    match = re.match(r"row-(\d+)__id-", custom_id)
    if not match:
        raise ValueError(f"Could not parse row number from custom_id: {custom_id}")
    return int(match.group(1))


def collect_batch_results(manifest: dict) -> pd.DataFrame:
    success_records = []
    error_records = []

    for batch_info in manifest.get("batches", []):
        output_path = batch_info.get("local_output_path")
        if output_path and os.path.exists(output_path):
            with open(output_path, "r", encoding="utf-8") as fh:
                for line in fh:
                    if not line.strip():
                        continue
                    payload = json.loads(line)
                    custom_id = payload.get("custom_id", "")
                    response_body = payload.get("response", {}).get("body", {})
                    row_number = parse_custom_id(custom_id)

                    try:
                        output_text = extract_output_text(response_body)
                        parsed = ClassificationResult.model_validate_json(output_text)
                        success_records.append(
                            {
                                "__row_number": row_number,
                                "Predicted_Taxonomy": parsed.Predicted_Taxonomy,
                                "Confidence_Score": parsed.Confidence_Score,
                                "Reasoning": parsed.Reasoning,
                                "Proposed_Taxonomy": parsed.Proposed_Taxonomy,
                            }
                        )
                    except Exception as exc:
                        error_records.append(
                            {
                                "__row_number": row_number,
                                "Predicted_Taxonomy": "Unknown > Unknown",
                                "Confidence_Score": "N/A",
                                "Reasoning": f"Result parsing error: {exc}",
                                "Proposed_Taxonomy": "",
                            }
                        )

        error_path = batch_info.get("local_error_path")
        if error_path and os.path.exists(error_path):
            with open(error_path, "r", encoding="utf-8") as fh:
                for line in fh:
                    if not line.strip():
                        continue
                    payload = json.loads(line)
                    custom_id = payload.get("custom_id", "")
                    row_number = parse_custom_id(custom_id) if custom_id else None
                    message = payload.get("error", {}).get("message", "Batch request failed")
                    error_records.append(
                        {
                            "__row_number": row_number,
                            "Predicted_Taxonomy": "Unknown > Unknown",
                            "Confidence_Score": "N/A",
                            "Reasoning": f"Batch error: {message}",
                            "Proposed_Taxonomy": "",
                        }
                    )

    results_df = pd.DataFrame(success_records + error_records)
    if results_df.empty:
        raise RuntimeError("No batch results were found. Check batch status and output files.")

    results_df = results_df.sort_values("__row_number").drop_duplicates(subset=["__row_number"], keep="first")
    return results_df


def format_and_save_excel(final_df: pd.DataFrame) -> None:
    OUTPUT_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_FILE_PATH, engine="openpyxl") as writer:
        final_df.to_excel(writer, index=False, sheet_name="Classifications")
        worksheet = writer.sheets["Classifications"]

        for idx, col in enumerate(final_df.columns):
            col_letter = get_column_letter(idx + 1)
            if col == "Reasoning":
                worksheet.column_dimensions[col_letter].width = 100
            else:
                max_len = max(final_df[col].astype(str).map(len).max(), len(col))
                worksheet.column_dimensions[col_letter].width = min(max_len + 2, 60)

            for cell in worksheet[col_letter]:
                if col == "Reasoning":
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")

        max_row = len(final_df) + 1
        max_col_letter = get_column_letter(len(final_df.columns))
        table_range = f"A1:{max_col_letter}{max_row}"
        tab = Table(displayName="ClassificationData", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        worksheet.add_table(tab)


def merge_results_to_excel(taxonomy_df: pd.DataFrame, items_df: pd.DataFrame, results_df: pd.DataFrame) -> pd.DataFrame:
    results_df[["Sub_Category", "Family_Category"]] = results_df["Predicted_Taxonomy"].str.split(
        " > ", n=1, expand=True
    )

    final_df = pd.merge(items_df, results_df, on="__row_number", how="left")

    lookup_df = taxonomy_df[
        ["Predicted_Taxonomy", "Major", "Major_Code", "Sub_Code", "Family_Code", "Full_Code"]
    ].drop_duplicates()
    lookup_df = lookup_df.rename(columns={"Major": "Major_Category"})

    final_df = pd.merge(final_df, lookup_df, on="Predicted_Taxonomy", how="left")
    final_df = final_df.drop(columns=["Predicted_Taxonomy"])

    final_df["Major_Category"] = final_df["Major_Category"].fillna("Unknown")
    final_df["Major_Code"] = final_df["Major_Code"].fillna("UNK")
    final_df["Sub_Code"] = final_df["Sub_Code"].fillna("UNK")
    final_df["Family_Code"] = final_df["Family_Code"].fillna("UNK")
    final_df["Full_Code"] = final_df["Full_Code"].fillna("UNK-00")
    final_df["Proposed_Taxonomy"] = final_df["Proposed_Taxonomy"].fillna("")

    original_cols = [col for col in items_df.columns if col != "__row_number"]
    final_df = final_df.drop(columns=["__row_number"])

    insert_idx = original_cols.index("Item Description") + 1 if "Item Description" in original_cols else len(original_cols)
    new_col_order = (
        original_cols[:insert_idx]
        + [
            "Major_Category",
            "Sub_Category",
            "Family_Category",
            "Confidence_Score",
            "Reasoning",
            "Proposed_Taxonomy",
        ]
        + original_cols[insert_idx:]
        + ["Major_Code", "Sub_Code", "Family_Code", "Full_Code"]
    )
    final_df = final_df[new_col_order]
    format_and_save_excel(final_df)
    return final_df


def do_prepare() -> None:
    _, items_df, _, developer_prompt = load_taxonomy_and_items()
    manifest = write_jsonl_shards(items_df, developer_prompt)
    print(f"Prepared {len(manifest['files'])} JSONL file(s) at {BATCH_INPUT_DIR}")
    for file_info in manifest["files"]:
        print(
            f"  - {file_info['path']} | requests={file_info['request_count']} | bytes={file_info['bytes']}"
        )


def do_submit(wait: bool) -> None:
    client = get_client()
    manifest = load_manifest()
    manifest = submit_batches(client, manifest)
    print("Submitted batch job(s):")
    for batch_info in manifest["batches"]:
        print(
            f"  - batch_id={batch_info['batch_id']} | status={batch_info['status']} | input={batch_info['local_input_path']}"
        )
    if wait:
        manifest = wait_for_batches(client, manifest)
        print("All batches reached a terminal status.")


def do_status() -> None:
    client = get_client()
    manifest = load_manifest()
    manifest = refresh_batch_statuses(client, manifest)
    for batch_info in manifest.get("batches", []):
        print(json.dumps(batch_info, indent=2))


def do_download_and_merge() -> None:
    client = get_client()
    manifest = load_manifest()
    manifest = refresh_batch_statuses(client, manifest)
    incomplete = [b for b in manifest.get("batches", []) if b.get("status") != "completed"]
    if incomplete:
        statuses = ", ".join(f"{b['batch_id']}={b['status']}" for b in incomplete)
        raise RuntimeError(f"Not all batches are completed yet: {statuses}")

    manifest = download_batch_files(client, manifest)
    taxonomy_df, items_df, _, _ = load_taxonomy_and_items()
    results_df = collect_batch_results(manifest)
    final_df = merge_results_to_excel(taxonomy_df, items_df, results_df)

    print(f"Downloaded outputs and wrote {len(final_df)} rows to {OUTPUT_FILE_PATH}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run Motion taxonomy classification via the OpenAI Batch API.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    subparsers.add_parser("prepare", help="Create batch-ready JSONL files from the source Excel workbook.")

    submit_parser = subparsers.add_parser("submit", help="Upload JSONL files and create batch jobs.")
    submit_parser.add_argument("--wait", action="store_true", help="Poll until all submitted batches reach a terminal state.")

    subparsers.add_parser("status", help="Refresh and print the current batch statuses.")
    subparsers.add_parser(
        "download", help="Download completed batch outputs, parse them, and create the final Excel output."
    )
    return parser


if __name__ == "__main__":
    ensure_dirs()
    parser = build_parser()
    args = parser.parse_args()

    if args.command == "prepare":
        do_prepare()
    elif args.command == "submit":
        do_submit(wait=args.wait)
    elif args.command == "status":
        do_status()
    elif args.command == "download":
        do_download_and_merge()
